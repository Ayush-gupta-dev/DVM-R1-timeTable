import openpyxl
from datetime import datetime

class ExcelTable:
    PASSWORD = "1234" 
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook[sheet_name]

    def save_workbook(self):
        self.workbook.save(self.file_path)

    def close_workbook(self):
        self.workbook.close()

    def format_date(self, date):
        return date.strftime("%d/%m/%Y") if isinstance(date, datetime) else date

    def parse_date(self, date):
        return datetime.strptime(date, "%d/%m/%Y") if not isinstance(date, datetime) else date


    def display_course_data(self, course_name):
        column_names = [cell.value for cell in self.sheet[1]]
        print(tuple(column_names))

        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 0 and row[0] and row[0].upper() == course_name.upper():
                formatted_date = self.format_date(row[2])
                print((row[0], row[1], formatted_date, row[3]))


    def add_course(self, course, course_code, exam_date, credit):
        for cell in self.sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
            if cell[0] is not None:
                if self.parse_date(cell[0]) == self.parse_date(exam_date):
                    print("there was a clash in examination, retry")
                    return
        row_number = 1
        while self.sheet.cell(row=row_number, column=1).value is not None:
            row_number += 1

        exam_date = self.parse_date(exam_date)

        self.sheet.cell(row=row_number, column=1, value=course)
        self.sheet.cell(row=row_number, column=2, value=course_code)
        self.sheet.cell(row=row_number, column=3, value=exam_date)
        self.sheet.cell(row=row_number, column=3).number_format = 'DD/MM/YYYY'
        self.sheet.cell(row=row_number, column=4, value=credit)

    def main():
            file_path = './DvmExcel.xlsx'
            sheet_name = 'Sheet1'

            course_manager = ExcelTable(file_path, sheet_name)
            while True:
                usr_choice=input("Print course info(1) or Add course with admin privilege(2) or exit(3) ")                                                                                                                                  
                try:
                    if(usr_choice=="1"):
                        while True:
                            print("Courses: ",end='')
                            for cell in course_manager.sheet.iter_rows(min_row=2, min_col=1, max_col=1,values_only=True):
                                if cell[0] is not None:
                                    print(cell[0],end=' | ')                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                        
                            course_name = input(" \n Enter the course name(or'e'=>exit): ")
                            course_manager.display_course_data(course_name)
                            if(course_name=="e"):
                                break
                    if(usr_choice=="2"):
                        password_attempt = input("Enter the password to add data: ")

                        if password_attempt != ExcelTable.PASSWORD:
                            print("Incorrect password. Write access denied. Viewing only.")
                        else:
                            print("Password accepted. Full access.")
                            new_course = input("Enter the new course name: ")
                            new_course_code = input("Enter the new course code: ")
                            new_exam_date = input("Enter the new exam date (DD/MM/YYYY): ")
                            new_credit = int(input("Enter the new credit: "))

                            course_manager.add_course(new_course, new_course_code, new_exam_date, new_credit)
                            course_manager.save_workbook()
                    if(usr_choice=="3"):
                        break

                except KeyError:
                    print(f"Sheet '{sheet_name}' not found.")

                finally:
                    course_manager.close_workbook()
            

